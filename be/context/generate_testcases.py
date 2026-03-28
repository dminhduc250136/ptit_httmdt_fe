"""
Generate LaptopVerse BE API Test Cases Excel
Based on: IMPLEMENTATION_PLAN.md, data.sql, data_extra.sql
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# ============================================================
# STYLE HELPERS
# ============================================================
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

HDR_FILL   = make_fill("1F4E79")  # Dark blue
TC_FILL    = make_fill("D6E4F0")  # Light blue (positive)
NEG_FILL   = make_fill("FDEBD0")  # Light orange (negative)
PASS_FILL  = make_fill("D5F5E3")  # Light green (pass)
FAIL_FILL  = make_fill("FADBD8")  # Light red (fail)
GROUP_FILL = make_fill("2E86C1")  # Medium blue for group rows

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
GRP_FONT   = Font(bold=True, color="FFFFFF", size=10)
BOLD       = Font(bold=True, size=9)
NORMAL     = Font(size=9)

thin = Side(style="thin", color="BDBDBD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

COLUMNS = [
    ("TC_ID",          12),
    ("Module",         14),
    ("Tên Test Case",  38),
    ("Loại",           10),
    ("Method",          8),
    ("Endpoint",       38),
    ("Headers",        28),
    ("Request Body / Params", 45),
    ("Pre-Condition",  35),
    ("Expected Status",12),
    ("Expected Response (mô tả)", 45),
    ("Kết quả",        12),
    ("Ghi chú",        25),
]

def create_sheet(name):
    ws = wb.create_sheet(title=name)
    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    return ws

def write_group(ws, row, group_name):
    ws.row_dimensions[row].height = 18
    cell = ws.cell(row=row, column=1, value=group_name)
    cell.fill = GROUP_FILL
    cell.font = GRP_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    return row + 1

def write_tc(ws, row, data: dict, is_positive=True):
    fill = TC_FILL if is_positive else NEG_FILL
    vals = [
        data.get("id",""),
        data.get("module",""),
        data.get("name",""),
        data.get("type", "Positive" if is_positive else "Negative"),
        data.get("method",""),
        data.get("endpoint",""),
        data.get("headers",""),
        data.get("body",""),
        data.get("pre",""),
        data.get("status",""),
        data.get("expected",""),
        "",   # Kết quả (để trống cho tester điền)
        data.get("note",""),
    ]
    for col_idx, val in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = fill
        cell.font = BOLD if col_idx == 1 else NORMAL
        cell.alignment = WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 55
    return row + 1

AUTH_HDR = "Content-Type: application/json"
TOKEN_HDR = "Authorization: Bearer <token>"
ADMIN_HDR = "Authorization: Bearer <admin_token>"

# ============================================================
# SHEET 1 — AUTH
# ============================================================
ws = create_sheet("01_Auth")
row = 2

row = write_group(ws, row, "POST /api/v1/auth/register — Đăng ký tài khoản")
row = write_tc(ws, row, {
    "id":"TC_AUTH_01", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/register",
    "name":"Đăng ký thành công với đầy đủ thông tin hợp lệ",
    "headers": AUTH_HDR,
    "body": '{\n  "fullName": "Nguyễn Test",\n  "email": "testuser@gmail.com",\n  "phone": "0987654321",\n  "password": "Password@123"\n}',
    "pre": "Email chưa tồn tại trong DB",
    "status": "201 Created",
    "expected": "Trả về id, email, fullName, phone, role='CUSTOMER', token (JWT hợp lệ)",
})
row = write_tc(ws, row, {
    "id":"TC_AUTH_02", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/register",
    "name":"Đăng ký thất bại — Email đã tồn tại (nguyenvana@gmail.com)",
    "headers": AUTH_HDR,
    "body": '{\n  "fullName": "Duplicate",\n  "email": "nguyenvana@gmail.com",\n  "phone": "0999888777",\n  "password": "Pass123"\n}',
    "pre": "Email nguyenvana@gmail.com đã có trong DB (data.sql)",
    "status": "409 Conflict",
    "expected": '{ "message": "Email đã được sử dụng" }',
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_AUTH_03", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/register",
    "name":"Đăng ký thất bại — Email sai định dạng",
    "headers": AUTH_HDR,
    "body": '{\n  "fullName": "Test",\n  "email": "not-an-email",\n  "phone": "0987654321",\n  "password": "Pass123"\n}',
    "pre": "",
    "status": "400 Bad Request",
    "expected": 'Validation error: email không đúng định dạng',
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_AUTH_04", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/register",
    "name":"Đăng ký thất bại — Phone sai định dạng (không bắt đầu bằng 0[3-9])",
    "headers": AUTH_HDR,
    "body": '{\n  "fullName": "Test",\n  "email": "new@test.com",\n  "phone": "123456789",\n  "password": "Pass123"\n}',
    "pre": "",
    "status": "400 Bad Request",
    "expected": "Validation error về số điện thoại",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_AUTH_05", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/register",
    "name":"Đăng ký thất bại — Thiếu field bắt buộc (email bị bỏ trống)",
    "headers": AUTH_HDR,
    "body": '{\n  "fullName": "Test",\n  "phone": "0987654321",\n  "password": "Pass123"\n}',
    "pre": "",
    "status": "400 Bad Request",
    "expected": "Validation error: email không được trống",
}, is_positive=False)

row = write_group(ws, row, "POST /api/v1/auth/login — Đăng nhập")
row = write_tc(ws, row, {
    "id":"TC_AUTH_06", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/login",
    "name":"Đăng nhập thành công với tài khoản CUSTOMER hợp lệ",
    "headers": AUTH_HDR,
    "body": '{\n  "email": "nguyenvana@gmail.com",\n  "password": "123456"\n}',
    "pre": "User tồn tại, mật khẩu BCrypt('123456') trong data.sql",
    "status": "200 OK",
    "expected": "Trả về token JWT, role='CUSTOMER', fullName='Nguyễn Văn A'",
})
row = write_tc(ws, row, {
    "id":"TC_AUTH_07", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/login",
    "name":"Đăng nhập thành công với tài khoản ADMIN",
    "headers": AUTH_HDR,
    "body": '{\n  "email": "admin@laptopverse.vn",\n  "password": "123456"\n}',
    "pre": "Admin account tồn tại trong data.sql",
    "status": "200 OK",
    "expected": "Trả về token JWT hợp lệ, role='ADMIN'",
})
row = write_tc(ws, row, {
    "id":"TC_AUTH_08", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/login",
    "name":"Đăng nhập thất bại — Mật khẩu sai",
    "headers": AUTH_HDR,
    "body": '{\n  "email": "nguyenvana@gmail.com",\n  "password": "wrongpassword"\n}',
    "pre": "User tồn tại",
    "status": "400 Bad Request",
    "expected": '{ "message": "Email hoặc mật khẩu không đúng" }',
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_AUTH_09", "module":"Auth", "method":"POST",
    "endpoint":"/api/v1/auth/login",
    "name":"Đăng nhập thất bại — Email không tồn tại",
    "headers": AUTH_HDR,
    "body": '{\n  "email": "ghost@notfound.com",\n  "password": "123456"\n}',
    "pre": "Email không tồn tại trong DB",
    "status": "400 Bad Request",
    "expected": '{ "message": "Email hoặc mật khẩu không đúng" }',
}, is_positive=False)

row = write_group(ws, row, "GET /api/v1/auth/me — Lấy profile")
row = write_tc(ws, row, {
    "id":"TC_AUTH_10", "module":"Auth", "method":"GET",
    "endpoint":"/api/v1/auth/me",
    "name":"Lấy profile thành công với token hợp lệ",
    "headers": TOKEN_HDR,
    "body": "(không có body)",
    "pre": "Đã đăng nhập, có JWT token hợp lệ",
    "status": "200 OK",
    "expected": "Trả về id, email, fullName, phone, avatarUrl, role",
})
row = write_tc(ws, row, {
    "id":"TC_AUTH_11", "module":"Auth", "method":"GET",
    "endpoint":"/api/v1/auth/me",
    "name":"Lấy profile thất bại — Không có token",
    "headers": "(không có Authorization header)",
    "body": "(không có body)",
    "pre": "",
    "status": "401 Unauthorized / 403 Forbidden",
    "expected": "Trả về lỗi xác thực",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_AUTH_12", "module":"Auth", "method":"GET",
    "endpoint":"/api/v1/auth/me",
    "name":"Lấy profile thất bại — Token hết hạn / không hợp lệ",
    "headers": "Authorization: Bearer invalid.token.here",
    "body": "",
    "pre": "Dùng token fake hoặc đã hết hạn",
    "status": "401 Unauthorized / 403 Forbidden",
    "expected": "Trả về lỗi xác thực",
}, is_positive=False)

# ============================================================
# SHEET 2 — CATALOG (Categories, Brands, Products)
# ============================================================
ws = create_sheet("02_Catalog")
row = 2

row = write_group(ws, row, "GET /api/v1/categories — Danh mục sản phẩm")
row = write_tc(ws, row, {
    "id":"TC_CAT_01", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/categories",
    "name":"Lấy danh sách tất cả danh mục active",
    "headers": "(không cần auth)",
    "body": "",
    "pre": "DB có 3 categories: Gaming, Văn phòng, Apple (data.sql)",
    "status": "200 OK",
    "expected": "Mảng 3 danh mục, mỗi item có: id, name, slug, icon, description, productCount. Sắp xếp theo displayOrder",
})
row = write_tc(ws, row, {
    "id":"TC_CAT_02", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/categories/gaming",
    "name":"Lấy chi tiết danh mục theo slug hợp lệ",
    "headers": "",
    "body": "",
    "pre": "Category 'gaming' tồn tại",
    "status": "200 OK",
    "expected": "Trả về danh mục Gaming với id=1, icon='Gamepad2', productCount > 0",
})
row = write_tc(ws, row, {
    "id":"TC_CAT_03", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/categories/khong-ton-tai",
    "name":"Lấy danh mục với slug không tồn tại",
    "headers": "",
    "body": "",
    "pre": "Slug 'khong-ton-tai' không có trong DB",
    "status": "404 Not Found",
    "expected": '{ "message": "..." }',
}, is_positive=False)

row = write_group(ws, row, "GET /api/v1/brands — Thương hiệu")
row = write_tc(ws, row, {
    "id":"TC_BRD_01", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/brands",
    "name":"Lấy danh sách tất cả brands active",
    "headers": "",
    "body": "",
    "pre": "DB có 8 brands: Acer, Apple, ASUS, Dell, HP, Lenovo, MSI, Razer (data.sql)",
    "status": "200 OK",
    "expected": "Mảng 8 brands, sắp xếp theo tên A-Z, có productCount",
})

row = write_group(ws, row, "GET /api/v1/products — Danh sách sản phẩm (có filter/sort/page)")
row = write_tc(ws, row, {
    "id":"TC_PRD_01", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Lấy trang đầu sản phẩm không filter (mặc định page=0, size=18)",
    "headers": "",
    "body": "?page=0&size=18",
    "pre": "DB có 18 sản phẩm active",
    "status": "200 OK",
    "expected": "Page object: content (18 items), totalElements=18, totalPages=1. Mỗi item có id, name, slug, brand, category, price, originalPrice, discountPercent, image, rating, reviewCount, soldCount, badge",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_02", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Filter theo category slug 'gaming'",
    "headers": "",
    "body": "?category=gaming",
    "pre": "Category gaming có: SP 1,3,7,9,12,14,17,18 → 8 sản phẩm gaming",
    "status": "200 OK",
    "expected": "Tất cả sản phẩm trong category gaming, brand đúng tên brand",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_03", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Filter theo brand 'apple'",
    "headers": "",
    "body": "?brand=apple",
    "pre": "Brand Apple có SP: 2,6,10,16 → 4 sản phẩm",
    "status": "200 OK",
    "expected": "4 sản phẩm Apple (MacBook Pro 16, MacBook Air 15, MacBook Pro 14, MacBook Air 13)",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_04", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Filter multi-brand (apple + msi)",
    "headers": "",
    "body": "?brand=apple&brand=msi",
    "pre": "",
    "status": "200 OK",
    "expected": "6 sản phẩm (4 Apple + 2 MSI)",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_05", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Filter theo khoảng giá 20M-40M VNĐ",
    "headers": "",
    "body": "?priceMin=20000000&priceMax=40000000",
    "pre": "",
    "status": "200 OK",
    "expected": "Chỉ trả về SP có price trong [20M, 40M]: SP 5(38.99M), 6(37.99M), 9(34.99M), 11(28.99M), 13(24.99M), 15(32.99M), 16(27.99M), 18(...).",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_06", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Filter theo badge 'Hot'",
    "headers": "",
    "body": "?badge=Hot",
    "pre": "SP có badge Hot: 1,2,5,6,14,16",
    "status": "200 OK",
    "expected": "Chỉ trả về 6 sản phẩm có badge=Hot",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_07", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Tìm kiếm theo keyword 'macbook'",
    "headers": "",
    "body": "?search=macbook",
    "pre": "",
    "status": "200 OK",
    "expected": "Trả về các SP có tên chứa 'macbook' (case-insensitive): 4 MacBook products",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_08", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Sắp xếp theo giá tăng dần (sortBy=price-asc)",
    "headers": "",
    "body": "?sortBy=price-asc",
    "pre": "",
    "status": "200 OK",
    "expected": "Sản phẩm đầu tiên là MacBook Air 13 M3 (27.99M) — giá thấp nhất",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_09", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Sắp xếp theo giá giảm dần (sortBy=price-desc)",
    "headers": "",
    "body": "?sortBy=price-desc",
    "pre": "",
    "status": "200 OK",
    "expected": "Sản phẩm đầu tiên là MSI Titan 18 HX 2024 (119.99M) — giá cao nhất",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_10", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Sắp xếp theo đánh giá cao nhất (sortBy=rating)",
    "headers": "",
    "body": "?sortBy=rating",
    "pre": "",
    "status": "200 OK",
    "expected": "SP đầu tiên có rating 4.9 (MacBook Pro 14 / MacBook Pro 16)",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_11", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Phân trang — page=0, size=5",
    "headers": "",
    "body": "?page=0&size=5",
    "pre": "",
    "status": "200 OK",
    "expected": "totalElements=18, totalPages=4, content có 5 items, number=0",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_12", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Phân trang — page vượt quá số trang thực",
    "headers": "",
    "body": "?page=99&size=18",
    "pre": "",
    "status": "200 OK",
    "expected": "totalElements=18, content=[] (mảng rỗng), number=99",
}, is_positive=False)

row = write_group(ws, row, "GET /api/v1/products/{id} — Chi tiết sản phẩm")
row = write_tc(ws, row, {
    "id":"TC_PRD_13", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products/1",
    "name":"Lấy chi tiết ASUS ROG Strix G16 2024 (id=1)",
    "headers": "",
    "body": "",
    "pre": "SP id=1 tồn tại (data.sql)",
    "status": "200 OK",
    "expected": "Có: id=1, name='ASUS ROG Strix G16 2024', brand={id,name,slug}, category={id,name,slug}, price=42990000, originalPrice=49990000, discountPercent=14, gallery (5 ảnh), detailedSpecs (12 specs), gifts (4 quà), rating=4.8, reviewCount=234, badge='Hot', isFlashSale=true",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_14", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products/2",
    "name":"Lấy chi tiết MacBook Pro 16 M3 Max (id=2)",
    "headers": "",
    "body": "",
    "pre": "SP id=2 tồn tại",
    "status": "200 OK",
    "expected": "price=89990000, brand='Apple', category='Apple', specs: cpu='Apple M3 Max',ram='48GB Unified', isFlashSale=true (đang trong flash sale data_extra.sql)",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_15", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products/10",
    "name":"Lấy chi tiết MacBook Pro 14 M3 Pro (id=10) — badge NULL",
    "headers": "",
    "body": "",
    "pre": "SP id=10 tồn tại, badge=NULL",
    "status": "200 OK",
    "expected": "badge trả về null hoặc không có field (Jackson @JsonInclude NON_NULL)",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_16", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products/9999",
    "name":"Lấy chi tiết SP với id không tồn tại",
    "headers": "",
    "body": "",
    "pre": "id=9999 không tồn tại",
    "status": "404 Not Found",
    "expected": '{ "status": 404, "message": "..." }',
}, is_positive=False)

row = write_group(ws, row, "GET /api/v1/products/{id}/related — Sản phẩm liên quan")
row = write_tc(ws, row, {
    "id":"TC_PRD_17", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products/1/related",
    "name":"Lấy sản phẩm liên quan của SP id=1 (gaming, ASUS)",
    "headers": "",
    "body": "",
    "pre": "SP id=1 tồn tại, category gaming",
    "status": "200 OK",
    "expected": "Danh sách SP cùng category gaming, không bao gồm SP id=1, tối đa 4-6 items",
})
row = write_tc(ws, row, {
    "id":"TC_PRD_18", "module":"Catalog", "method":"GET",
    "endpoint":"/api/v1/products/9999/related",
    "name":"Lấy related cho SP không tồn tại",
    "headers": "",
    "body": "",
    "pre": "",
    "status": "404 Not Found",
    "expected": "Resource not found error",
}, is_positive=False)

# ============================================================
# SHEET 3 — FLASH SALE
# ============================================================
ws = create_sheet("03_FlashSale")
row = 2

row = write_group(ws, row, "GET /api/v1/flash-sales/active — Flash sale đang diễn ra")
row = write_tc(ws, row, {
    "id":"TC_FS_01", "module":"FlashSale", "method":"GET",
    "endpoint":"/api/v1/flash-sales/active",
    "name":"Lấy flash sale đang active (có dữ liệu)",
    "headers": "",
    "body": "",
    "pre": "Flash sale 'Flash Sale Cuối Tuần' đang active (NOW + 8h trong data_extra.sql). 5 sản phẩm: id=1,2,3,4,9",
    "status": "200 OK",
    "expected": "Trả về FlashSaleDTO: id=1, title='Flash Sale Cuối Tuần', startTime, endTime, products (5 items). Mỗi item có flashSalePrice, stockLimit, flashSaleSold, discountPercent",
})
row = write_tc(ws, row, {
    "id":"TC_FS_02", "module":"FlashSale", "method":"GET",
    "endpoint":"/api/v1/flash-sales/active",
    "name":"Không có flash sale active",
    "headers": "",
    "body": "",
    "pre": "Tất cả flash sales đã hết hạn hoặc chưa bắt đầu (cập nhật DB: set is_active=false)",
    "status": "204 No Content",
    "expected": "Trả về 204 không có body",
}, is_positive=False)

# ============================================================
# SHEET 4 — CART
# ============================================================
ws = create_sheet("04_Cart")
row = 2

row = write_group(ws, row, "GET /api/v1/cart — Xem giỏ hàng")
row = write_tc(ws, row, {
    "id":"TC_CART_01", "module":"Cart", "method":"GET",
    "endpoint":"/api/v1/cart",
    "name":"Xem giỏ hàng (giỏ trống)",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "User đã đăng nhập, chưa thêm gì vào giỏ",
    "status": "200 OK",
    "expected": "{ items: [], totalItems: 0, totalPrice: 0 }",
})
row = write_tc(ws, row, {
    "id":"TC_CART_02", "module":"Cart", "method":"GET",
    "endpoint":"/api/v1/cart",
    "name":"Xem giỏ hàng (có sản phẩm)",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "User đã thêm SP id=1 vào giỏ",
    "status": "200 OK",
    "expected": "items có 1 entry với product.id=1, quantity=1. totalPrice = 42990000. CartProductDTO có id, name, slug, image, price, originalPrice",
})
row = write_tc(ws, row, {
    "id":"TC_CART_03", "module":"Cart", "method":"GET",
    "endpoint":"/api/v1/cart",
    "name":"Xem giỏ hàng — Không có token",
    "headers": "(không có Authorization)",
    "body": "",
    "pre": "",
    "status": "401/403",
    "expected": "Lỗi unauthorized",
}, is_positive=False)

row = write_group(ws, row, "POST /api/v1/cart — Thêm sản phẩm vào giỏ")
row = write_tc(ws, row, {
    "id":"TC_CART_04", "module":"Cart", "method":"POST",
    "endpoint":"/api/v1/cart",
    "name":"Thêm SP mới vào giỏ (SP id=1 chưa trong giỏ)",
    "headers": TOKEN_HDR,
    "body": '{ "productId": 1 }',
    "pre": "User đăng nhập, SP id=1 active, giỏ chưa có SP này",
    "status": "200 OK",
    "expected": "Giỏ hàng trả về có SP id=1, quantity=1",
})
row = write_tc(ws, row, {
    "id":"TC_CART_05", "module":"Cart", "method":"POST",
    "endpoint":"/api/v1/cart",
    "name":"Thêm SP đã có trong giỏ → tăng quantity lên 1",
    "headers": TOKEN_HDR,
    "body": '{ "productId": 1 }',
    "pre": "SP id=1 đã trong giỏ với quantity=1",
    "status": "200 OK",
    "expected": "SP id=1 trong giỏ có quantity=2",
})
row = write_tc(ws, row, {
    "id":"TC_CART_06", "module":"Cart", "method":"POST",
    "endpoint":"/api/v1/cart",
    "name":"Thêm SP khi quantity đã đạt tối đa (10) → lỗi",
    "headers": TOKEN_HDR,
    "body": '{ "productId": 1 }',
    "pre": "SP id=1 đã trong giỏ với quantity=10",
    "status": "400 Bad Request",
    "expected": '{ "message": "Số lượng tối đa cho mỗi sản phẩm là 10" }',
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_CART_07", "module":"Cart", "method":"POST",
    "endpoint":"/api/v1/cart",
    "name":"Thêm SP không tồn tại vào giỏ",
    "headers": TOKEN_HDR,
    "body": '{ "productId": 9999 }',
    "pre": "",
    "status": "404 Not Found",
    "expected": "Sản phẩm không tồn tại",
}, is_positive=False)

row = write_group(ws, row, "PUT /api/v1/cart/{productId} — Cập nhật số lượng")
row = write_tc(ws, row, {
    "id":"TC_CART_08", "module":"Cart", "method":"PUT",
    "endpoint":"/api/v1/cart/1",
    "name":"Cập nhật quantity SP id=1 thành 3",
    "headers": TOKEN_HDR,
    "body": '{ "quantity": 3 }',
    "pre": "SP id=1 đã trong giỏ",
    "status": "200 OK",
    "expected": "SP id=1 có quantity=3, totalPrice = 3 × 42990000 = 128970000",
})
row = write_tc(ws, row, {
    "id":"TC_CART_09", "module":"Cart", "method":"PUT",
    "endpoint":"/api/v1/cart/1",
    "name":"Cập nhật quantity = 0 → xóa khỏi giỏ",
    "headers": TOKEN_HDR,
    "body": '{ "quantity": 0 }',
    "pre": "SP id=1 đã trong giỏ",
    "status": "200 OK",
    "expected": "SP id=1 bị xóa khỏi giỏ, items rỗng hoặc không chứa SP id=1",
})
row = write_tc(ws, row, {
    "id":"TC_CART_10", "module":"Cart", "method":"PUT",
    "endpoint":"/api/v1/cart/1",
    "name":"Cập nhật quantity > 10 → lỗi",
    "headers": TOKEN_HDR,
    "body": '{ "quantity": 11 }',
    "pre": "",
    "status": "400 Bad Request",
    "expected": "Validation error: quantity tối đa 10",
}, is_positive=False)

row = write_group(ws, row, "DELETE /api/v1/cart/{productId} — Xóa 1 item")
row = write_tc(ws, row, {
    "id":"TC_CART_11", "module":"Cart", "method":"DELETE",
    "endpoint":"/api/v1/cart/1",
    "name":"Xóa SP id=1 khỏi giỏ hàng",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "SP id=1 trong giỏ",
    "status": "204 No Content",
    "expected": "Không có body. SP id=1 không còn trong giỏ khi GET /cart",
})

row = write_group(ws, row, "DELETE /api/v1/cart — Xóa toàn bộ giỏ")
row = write_tc(ws, row, {
    "id":"TC_CART_12", "module":"Cart", "method":"DELETE",
    "endpoint":"/api/v1/cart",
    "name":"Xóa toàn bộ giỏ hàng",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "Giỏ hàng có nhiều item",
    "status": "204 No Content",
    "expected": "Sau khi xóa, GET /cart trả về { items: [], totalItems: 0, totalPrice: 0 }",
})

# ============================================================
# SHEET 5 — ORDERS
# ============================================================
ws = create_sheet("05_Orders")
row = 2

row = write_group(ws, row, "POST /api/v1/orders — Tạo đơn hàng")
row = write_tc(ws, row, {
    "id":"TC_ORD_01", "module":"Order", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"Tạo đơn hàng thành công với địa chỉ mới (COD)",
    "headers": TOKEN_HDR,
    "body": '{\n  "items": [{"productId":1,"quantity":1}],\n  "shippingAddress": {\n    "fullName":"Nguyễn Văn A",\n    "phone":"0912345678",\n    "province":"TPHCM",\n    "district":"Q7",\n    "address":"123 Nguyễn Văn Linh"\n  },\n  "paymentMethod": "cod"\n}',
    "pre": "SP id=1 active, stockQuantity=50",
    "status": "201 Created",
    "expected": "OrderDTO: orderCode bắt đầu 'ORD-', orderStatus='PENDING', paymentStatus='PENDING', subtotal=42990000, shippingFee=30000 (dưới 10M), total=43020000",
})
row = write_tc(ws, row, {
    "id":"TC_ORD_02", "module":"Order", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"Tạo đơn hàng free ship — subtotal >= 10.000.000đ",
    "headers": TOKEN_HDR,
    "body": '{\n  "items": [{"productId":2,"quantity":1}],\n  "shippingAddress": {\n    "fullName":"Nguyễn Văn A","phone":"0912345678",\n    "province":"TPHCM","district":"Q7","address":"123 Test"\n  },\n  "paymentMethod": "cod"\n}',
    "pre": "SP id=2 giá 89990000 > 10M",
    "status": "201 Created",
    "expected": "shippingFee=0 (miễn phí ship), total = subtotal",
})
row = write_tc(ws, row, {
    "id":"TC_ORD_03", "module":"Order", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"Tạo đơn hàng với địa chỉ đã lưu (addressId)",
    "headers": TOKEN_HDR,
    "body": '{\n  "items": [{"productId":5,"quantity":1}],\n  "shippingAddress": {"id": 1},\n  "paymentMethod": "vnpay"\n}',
    "pre": "User id=2 có address id=1 trong DB (data_extra.sql)",
    "status": "201 Created",
    "expected": "Đơn hàng tạo thành công với địa chỉ từ shippingAddress id=1",
})
row = write_tc(ws, row, {
    "id":"TC_ORD_04", "module":"Order", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"Tạo đơn hàng nhiều sản phẩm",
    "headers": TOKEN_HDR,
    "body": '{\n  "items": [\n    {"productId":11,"quantity":2},\n    {"productId":13,"quantity":1}\n  ],\n  "shippingAddress": {\n    "fullName":"Test","phone":"0912345678",\n    "province":"HN","district":"BD","address":"45 Test Street"\n  },\n  "paymentMethod": "cod"\n}',
    "pre": "SP 11 (28990000×2=57980000), SP 13 (24990000×1) → subtotal=82970000>=10M → free ship",
    "status": "201 Created",
    "expected": "2 order items, shippingFee=0, subtotal=82970000, total=82970000",
})
row = write_tc(ws, row, {
    "id":"TC_ORD_05", "module":"Order", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"Tạo đơn thất bại — Thiếu items",
    "headers": TOKEN_HDR,
    "body": '{ "items": [], "shippingAddress": {...}, "paymentMethod": "cod" }',
    "pre": "",
    "status": "400 Bad Request",
    "expected": "Validation error: Danh sách sản phẩm không được trống",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_ORD_06", "module":"Order", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"Tạo đơn thất bại — SP không tồn tại",
    "headers": TOKEN_HDR,
    "body": '{ "items": [{"productId":9999,"quantity":1}], "shippingAddress": {...}, "paymentMethod": "cod" }',
    "pre": "",
    "status": "404 Not Found",
    "expected": "Sản phẩm không tồn tại",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_ORD_07", "module":"Order", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"Tạo đơn thất bại — Phương thức thanh toán không hợp lệ",
    "headers": TOKEN_HDR,
    "body": '{ "items": [{"productId":1,"quantity":1}], "shippingAddress": {...}, "paymentMethod": "bitcoin" }',
    "pre": "",
    "status": "400/500",
    "expected": "Lỗi enum conversion hoặc bad request",
}, is_positive=False)

row = write_group(ws, row, "GET /api/v1/orders — Danh sách đơn hàng của user")
row = write_tc(ws, row, {
    "id":"TC_ORD_08", "module":"Order", "method":"GET",
    "endpoint":"/api/v1/orders",
    "name":"Xem tất cả đơn hàng của user",
    "headers": TOKEN_HDR,
    "body": "?page=0&size=10",
    "pre": "User đã có đơn hàng",
    "status": "200 OK",
    "expected": "Page object với danh sách OrderListDTO: id, orderCode, total, orderStatus, paymentMethod, createdAt, itemCount",
})
row = write_tc(ws, row, {
    "id":"TC_ORD_09", "module":"Order", "method":"GET",
    "endpoint":"/api/v1/orders",
    "name":"Filter đơn hàng theo status PENDING",
    "headers": TOKEN_HDR,
    "body": "?status=PENDING",
    "pre": "User có đơn PENDING",
    "status": "200 OK",
    "expected": "Chỉ trả về đơn có orderStatus=PENDING",
})

row = write_group(ws, row, "GET /api/v1/orders/{orderCode} — Chi tiết đơn hàng")
row = write_tc(ws, row, {
    "id":"TC_ORD_10", "module":"Order", "method":"GET",
    "endpoint":"/api/v1/orders/ORD-XXXXXXXX",
    "name":"Xem chi tiết đơn hàng của chính mình",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "Dùng orderCode từ TC_ORD_01, đúng user",
    "status": "200 OK",
    "expected": "OrderDTO đầy đủ: items, shippingAddress, subtotal, shippingFee, total, paymentMethod, orderStatus",
})
row = write_tc(ws, row, {
    "id":"TC_ORD_11", "module":"Order", "method":"GET",
    "endpoint":"/api/v1/orders/ORD-XXXXXXXX",
    "name":"Xem đơn hàng của người khác → không tìm thấy",
    "headers": "Authorization: Bearer <token_of_user_B>",
    "body": "",
    "pre": "orderCode thuộc user A, nhưng đang dùng token user B",
    "status": "404 Not Found",
    "expected": "Đơn hàng không tồn tại (security: không lộ thông tin)",
}, is_positive=False)

row = write_group(ws, row, "PATCH /api/v1/orders/{orderCode}/cancel — Hủy đơn")
row = write_tc(ws, row, {
    "id":"TC_ORD_12", "module":"Order", "method":"PATCH",
    "endpoint":"/api/v1/orders/ORD-XXXXXXXX/cancel",
    "name":"Hủy đơn hàng PENDING thành công",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "Đơn hàng đang ở trạng thái PENDING",
    "status": "200 OK",
    "expected": "orderStatus='CANCELLED'. Stock sản phẩm được cộng lại",
})
row = write_tc(ws, row, {
    "id":"TC_ORD_13", "module":"Order", "method":"PATCH",
    "endpoint":"/api/v1/orders/ORD-XXXXXXXX/cancel",
    "name":"Hủy đơn hàng đang SHIPPING → lỗi",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "Đơn hàng ở trạng thái SHIPPING (update thủ công trong DB)",
    "status": "400 Bad Request",
    "expected": '{ "message": "Không thể hủy đơn hàng ở trạng thái SHIPPING" }',
}, is_positive=False)

# ============================================================
# SHEET 6 — REVIEWS
# ============================================================
ws = create_sheet("06_Reviews")
row = 2

row = write_group(ws, row, "GET /api/v1/reviews/product/{productId} — Lấy danh sách review")
row = write_tc(ws, row, {
    "id":"TC_REV_01", "module":"Review", "method":"GET",
    "endpoint":"/api/v1/reviews/product/1",
    "name":"Lấy reviews của SP id=1 (ASUS ROG)",
    "headers": "",
    "body": "?page=0&size=10",
    "pre": "SP id=1 có 3 reviews trong data_extra.sql từ user 2,3,4",
    "status": "200 OK",
    "expected": "Page reviews: 3 items, mỗi item có id, rating, title, content, helpfulCount, isVerified, createdAt, userName, userAvatar",
})
row = write_tc(ws, row, {
    "id":"TC_REV_02", "module":"Review", "method":"GET",
    "endpoint":"/api/v1/reviews/product/16",
    "name":"Lấy reviews của MacBook Air 13 M3 (id=16)",
    "headers": "",
    "body": "?page=0&size=10",
    "pre": "SP id=16 có 3 reviews: user2(5*), user3(5*), user5(4*)",
    "status": "200 OK",
    "expected": "3 reviews, sắp xếp mới nhất trước, isVerified của từng user",
})
row = write_tc(ws, row, {
    "id":"TC_REV_03", "module":"Review", "method":"GET",
    "endpoint":"/api/v1/reviews/product/9999",
    "name":"Lấy reviews SP không tồn tại",
    "headers": "",
    "body": "",
    "pre": "",
    "status": "404 Not Found",
    "expected": "Sản phẩm không tồn tại",
}, is_positive=False)

row = write_group(ws, row, "GET /api/v1/reviews/product/{productId}/summary — Tổng hợp review")
row = write_tc(ws, row, {
    "id":"TC_REV_04", "module":"Review", "method":"GET",
    "endpoint":"/api/v1/reviews/product/1/summary",
    "name":"Lấy tóm tắt review SP id=1",
    "headers": "",
    "body": "",
    "pre": "SP id=1 có 3 reviews: 5,5,4 (avg≈4.67, totalReviews=3)",
    "status": "200 OK",
    "expected": "avgRating~4.67, totalReviews=3, distribution: {4:1, 5:2, 1:0, 2:0, 3:0}",
})
row = write_tc(ws, row, {
    "id":"TC_REV_05", "module":"Review", "method":"GET",
    "endpoint":"/api/v1/reviews/product/16/summary",
    "name":"Lấy tóm tắt review SP id=16 (MacBook Air 13)",
    "headers": "",
    "body": "",
    "pre": "SP id=16 có 3 reviews: 5,5,4 (avg≈4.67)",
    "status": "200 OK",
    "expected": "avgRating≈4.67, totalReviews=3, distribution có {4:1, 5:2}",
})

row = write_group(ws, row, "POST /api/v1/reviews/product/{productId} — Tạo review")
row = write_tc(ws, row, {
    "id":"TC_REV_06", "module":"Review", "method":"POST",
    "endpoint":"/api/v1/reviews/product/3",
    "name":"Tạo review hợp lệ cho SP mà user chưa review (SP id=3, user=nguyenvana)",
    "headers": TOKEN_HDR,
    "body": '{\n  "rating": 5,\n  "title": "Siêu phẩm MSI Titan!",\n  "content": "Hiệu năng không thể tin được, chơi game 4K mượt tuyệt đối."\n}',
    "pre": "User 2 (nguyenvana) chưa review SP id=3, đã mua SP này",
    "status": "201 Created",
    "expected": "ReviewDTO với isVerified=true (đã mua hàng), rating=5",
})
row = write_tc(ws, row, {
    "id":"TC_REV_07", "module":"Review", "method":"POST",
    "endpoint":"/api/v1/reviews/product/5",
    "name":"Tạo review — chưa mua hàng, isVerified=false",
    "headers": TOKEN_HDR,
    "body": '{\n  "rating": 4,\n  "title": "Máy tốt",\n  "content": "Nghe nói ThinkPad rất bền, đang định mua."\n}',
    "pre": "User chưa mua SP id=5",
    "status": "201 Created",
    "expected": "ReviewDTO với isVerified=false",
})
row = write_tc(ws, row, {
    "id":"TC_REV_08", "module":"Review", "method":"POST",
    "endpoint":"/api/v1/reviews/product/1",
    "name":"Tạo review thất bại — User đã review SP này (user2 đã review SP1)",
    "headers": TOKEN_HDR,
    "body": '{\n  "rating": 3,\n  "title": "Lần 2",\n  "content": "Review lại lần 2"\n}',
    "pre": "User id=2 đã có review cho SP id=1 (data_extra.sql)",
    "status": "409 Conflict",
    "expected": '{ "message": "Bạn đã đánh giá sản phẩm này" }',
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_REV_09", "module":"Review", "method":"POST",
    "endpoint":"/api/v1/reviews/product/1",
    "name":"Tạo review thất bại — rating < 1",
    "headers": TOKEN_HDR,
    "body": '{\n  "rating": 0,\n  "title": "Test",\n  "content": "Test content"\n}',
    "pre": "",
    "status": "400 Bad Request",
    "expected": "Validation error: rating phải >= 1",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_REV_10", "module":"Review", "method":"POST",
    "endpoint":"/api/v1/reviews/product/1",
    "name":"Tạo review không có token → lỗi auth",
    "headers": "(không có Authorization)",
    "body": '{ "rating": 5, "title": "T", "content": "C" }',
    "pre": "",
    "status": "401/403",
    "expected": "Unauthorized",
}, is_positive=False)

# ============================================================
# SHEET 7 — SHIPPING ADDRESSES
# ============================================================
ws = create_sheet("07_Addresses")
row = 2

row = write_group(ws, row, "GET /api/v1/addresses — Danh sách địa chỉ")
row = write_tc(ws, row, {
    "id":"TC_ADDR_01", "module":"Address", "method":"GET",
    "endpoint":"/api/v1/addresses",
    "name":"Lấy danh sách địa chỉ của user (có dữ liệu)",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "User id=2 có 1 địa chỉ mặc định (data_extra.sql)",
    "status": "200 OK",
    "expected": "Mảng 1 địa chỉ: fullName='Nguyễn Văn A', province='TP. Hồ Chí Minh', district='Quận 7', isDefault=true. Sắp xếp: default trước",
})
row = write_tc(ws, row, {
    "id":"TC_ADDR_02", "module":"Address", "method":"GET",
    "endpoint":"/api/v1/addresses",
    "name":"Lấy danh sách địa chỉ khi chưa có địa chỉ nào",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "User mới tạo, không có address",
    "status": "200 OK",
    "expected": "Mảng rỗng []",
})

row = write_group(ws, row, "POST /api/v1/addresses — Tạo địa chỉ mới")
row = write_tc(ws, row, {
    "id":"TC_ADDR_03", "module":"Address", "method":"POST",
    "endpoint":"/api/v1/addresses",
    "name":"Tạo địa chỉ mới hợp lệ",
    "headers": TOKEN_HDR,
    "body": '{\n  "fullName": "Nguyễn Văn A",\n  "phone": "0912345678",\n  "email": "a@test.com",\n  "province": "Hà Nội",\n  "district": "Hoàn Kiếm",\n  "address": "12 Hàng Bài, Phường Hàng Bài",\n  "isDefault": true\n}',
    "pre": "User đã đăng nhập",
    "status": "201 Created",
    "expected": "ShippingAddressDTO với id mới, isDefault=true. Các địa chỉ cũ isDefault bị chuyển thành false",
})
row = write_tc(ws, row, {
    "id":"TC_ADDR_04", "module":"Address", "method":"POST",
    "endpoint":"/api/v1/addresses",
    "name":"Tạo địa chỉ thất bại — Thiếu fullName",
    "headers": TOKEN_HDR,
    "body": '{\n  "phone": "0912345678",\n  "province": "HN",\n  "district": "Q1",\n  "address": "12 Test Street (dài hơn 10 ký tự)"\n}',
    "pre": "",
    "status": "400 Bad Request",
    "expected": "Validation error: fullName",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_ADDR_05", "module":"Address", "method":"POST",
    "endpoint":"/api/v1/addresses",
    "name":"Tạo địa chỉ thất bại — Phone sai định dạng",
    "headers": TOKEN_HDR,
    "body": '{\n  "fullName": "Test User",\n  "phone": "123",\n  "province": "HN",\n  "district": "Q1",\n  "address": "12 Test Street Address"\n}',
    "pre": "",
    "status": "400 Bad Request",
    "expected": "Validation error: phone",
}, is_positive=False)

row = write_group(ws, row, "PUT /api/v1/addresses/{id} — Cập nhật địa chỉ")
row = write_tc(ws, row, {
    "id":"TC_ADDR_06", "module":"Address", "method":"PUT",
    "endpoint":"/api/v1/addresses/1",
    "name":"Cập nhật địa chỉ id=1 của user hợp lệ",
    "headers": TOKEN_HDR,
    "body": '{\n  "fullName": "Nguyễn Văn A Updated",\n  "phone": "0912345678",\n  "province": "TP. Hồ Chí Minh",\n  "district": "Quận 1",\n  "address": "456 Lê Lợi, Phường Bến Nghé",\n  "isDefault": true\n}',
    "pre": "Address id=1 thuộc user id=2",
    "status": "200 OK",
    "expected": "fullName='Nguyễn Văn A Updated', district='Quận 1'",
})
row = write_tc(ws, row, {
    "id":"TC_ADDR_07", "module":"Address", "method":"PUT",
    "endpoint":"/api/v1/addresses/9999",
    "name":"Cập nhật địa chỉ không tồn tại",
    "headers": TOKEN_HDR,
    "body": '{ "fullName": "Test", "phone": "0912345678", "province": "HN", "district": "Q1", "address": "12 Test Street" }',
    "pre": "",
    "status": "404 Not Found",
    "expected": "Địa chỉ không tồn tại",
}, is_positive=False)

row = write_group(ws, row, "DELETE /api/v1/addresses/{id} — Xóa địa chỉ")
row = write_tc(ws, row, {
    "id":"TC_ADDR_08", "module":"Address", "method":"DELETE",
    "endpoint":"/api/v1/addresses/1",
    "name":"Xóa địa chỉ id=1 thành công",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "Address id=1 thuộc user id=2",
    "status": "204 No Content",
    "expected": "Không còn address id=1 trong GET /addresses",
})
row = write_tc(ws, row, {
    "id":"TC_ADDR_09", "module":"Address", "method":"DELETE",
    "endpoint":"/api/v1/addresses/3",
    "name":"Xóa địa chỉ của user khác → không tìm thấy",
    "headers": "Authorization: Bearer <token_user_2>",
    "body": "",
    "pre": "Address id=3 thuộc user id=4, không thuộc user id=2",
    "status": "404 Not Found",
    "expected": "Địa chỉ không tồn tại",
}, is_positive=False)

# ============================================================
# SHEET 8 — SECURITY (Cross-cutting)
# ============================================================
ws = create_sheet("08_Security")
row = 2

row = write_group(ws, row, "Kiểm tra xác thực & phân quyền")
row = write_tc(ws, row, {
    "id":"TC_SEC_01", "module":"Security", "method":"GET",
    "endpoint":"/api/v1/cart",
    "name":"Truy cập endpoint yêu cầu auth không có token",
    "headers": "(không có Authorization)",
    "body": "",
    "pre": "",
    "status": "401/403",
    "expected": "Lỗi xác thực, không trả về dữ liệu",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_SEC_02", "module":"Security", "method":"GET",
    "endpoint":"/api/v1/cart",
    "name":"Truy cập với token hết hạn",
    "headers": "Authorization: Bearer <expired_token>",
    "body": "",
    "pre": "Dùng token đã hết hạn (sau 24h)",
    "status": "401/403",
    "expected": "Lỗi token hết hạn / invalid",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_SEC_03", "module":"Security", "method":"GET",
    "endpoint":"/api/v1/admin/products",
    "name":"Truy cập admin endpoint với token CUSTOMER",
    "headers": TOKEN_HDR + " (CUSTOMER role)",
    "body": "",
    "pre": "Đăng nhập với nguyenvana@gmail.com (CUSTOMER)",
    "status": "403 Forbidden",
    "expected": "Bạn không có quyền truy cập tài nguyên này",
}, is_positive=False)
row = write_tc(ws, row, {
    "id":"TC_SEC_04", "module":"Security", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Truy cập public endpoint không cần token",
    "headers": "(không có Authorization)",
    "body": "",
    "pre": "",
    "status": "200 OK",
    "expected": "Danh sách sản phẩm trả về bình thường",
})
row = write_tc(ws, row, {
    "id":"TC_SEC_05", "module":"Security", "method":"GET",
    "endpoint":"/api/v1/products",
    "name":"Truy cập public endpoint với token CUSTOMER hợp lệ",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "",
    "status": "200 OK",
    "expected": "Danh sách sản phẩm — token không ảnh hưởng đến public API",
})
row = write_tc(ws, row, {
    "id":"TC_SEC_06", "module":"Security", "method":"GET",
    "endpoint":"/swagger-ui/index.html",
    "name":"Truy cập Swagger UI (public)",
    "headers": "",
    "body": "",
    "pre": "",
    "status": "200 OK",
    "expected": "Swagger UI hiển thị",
})

# ============================================================
# SHEET 9 — INTEGRATION (Flow tests)
# ============================================================
ws = create_sheet("09_Integration_Flow")
row = 2

row = write_group(ws, row, "Flow 1: Đăng ký → Đăng nhập → Mua hàng → Review")
row = write_tc(ws, row, {
    "id":"TC_FLOW_01", "module":"Integration", "method":"POST",
    "endpoint":"/api/v1/auth/register",
    "name":"[FLOW1-Step1] Đăng ký tài khoản mới",
    "headers": AUTH_HDR,
    "body": '{ "fullName":"Flow User", "email":"flow@test.com", "phone":"0977777777", "password":"Test@123" }',
    "pre": "Email chưa tồn tại",
    "status": "201",
    "expected": "Nhận token JWT",
    "note": "Lưu token cho các step tiếp theo"
})
row = write_tc(ws, row, {
    "id":"TC_FLOW_02", "module":"Integration", "method":"POST",
    "endpoint":"/api/v1/cart",
    "name":"[FLOW1-Step2] Thêm MacBook Air 13 (id=16) vào giỏ",
    "headers": "Authorization: Bearer <token từ FLOW1-Step1>",
    "body": '{ "productId": 16 }',
    "pre": "Đã có token từ step 1",
    "status": "200",
    "expected": "Giỏ hàng có SP id=16, quantity=1",
})
row = write_tc(ws, row, {
    "id":"TC_FLOW_03", "module":"Integration", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"[FLOW1-Step3] Đặt hàng SP id=16",
    "headers": "Authorization: Bearer <token từ FLOW1-Step1>",
    "body": '{ "items":[{"productId":16,"quantity":1}], "shippingAddress":{"fullName":"Flow User","phone":"0977777777","province":"HCM","district":"Q3","address":"123 Flow Street Test"}, "paymentMethod":"cod" }',
    "pre": "Có token hợp lệ",
    "status": "201",
    "expected": "Nhận orderCode, orderStatus=PENDING",
    "note": "Lưu orderCode để test bước tiếp"
})
row = write_tc(ws, row, {
    "id":"TC_FLOW_04", "module":"Integration", "method":"GET",
    "endpoint":"/api/v1/cart",
    "name":"[FLOW1-Step4] Kiểm tra giỏ hàng sau đặt hàng — SP đã mua bị xóa",
    "headers": "Authorization: Bearer <token>",
    "body": "",
    "pre": "Đã đặt hàng SP id=16 ở step 3",
    "status": "200",
    "expected": "Giỏ hàng rỗng (SP id=16 bị xóa sau khi order)",
})
row = write_tc(ws, row, {
    "id":"TC_FLOW_05", "module":"Integration", "method":"POST",
    "endpoint":"/api/v1/reviews/product/16",
    "name":"[FLOW1-Step5] Review SP sau khi đặt hàng → isVerified=true",
    "headers": "Authorization: Bearer <token>",
    "body": '{ "rating": 5, "title": "Tuyệt vời!", "content": "MacBook Air 13 M3 rất nhanh và nhẹ, pin trâu!" }',
    "pre": "Đã đặt hàng SP id=16 (step 3)",
    "status": "201",
    "expected": "isVerified=true (đã mua hàng). SP id=16 rating tự cập nhật",
})

row = write_group(ws, row, "Flow 2: Hủy đơn hàng + Kiểm tra stock hoàn lại")
row = write_tc(ws, row, {
    "id":"TC_FLOW_06", "module":"Integration", "method":"GET",
    "endpoint":"/api/v1/products/9",
    "name":"[FLOW2-Step1] Kiểm tra stock SP id=9 trước khi đặt",
    "headers": "",
    "body": "",
    "pre": "SP id=9 stock=55 (data.sql)",
    "status": "200",
    "expected": "stockQuantity=55",
    "note": "Ghi nhớ stockQuantity ban đầu"
})
row = write_tc(ws, row, {
    "id":"TC_FLOW_07", "module":"Integration", "method":"POST",
    "endpoint":"/api/v1/orders",
    "name":"[FLOW2-Step2] Đặt hàng 2 SP id=9",
    "headers": TOKEN_HDR,
    "body": '{ "items":[{"productId":9,"quantity":2}], "shippingAddress":{"fullName":"Test","phone":"0912345678","province":"HCM","district":"Q1","address":"456 Test Address 123"}, "paymentMethod":"cod" }',
    "pre": "",
    "status": "201",
    "expected": "Đơn hàng tạo thành công. stock SP id=9 giảm còn 53",
})
row = write_tc(ws, row, {
    "id":"TC_FLOW_08", "module":"Integration", "method":"PATCH",
    "endpoint":"/api/v1/orders/<orderCode>/cancel",
    "name":"[FLOW2-Step3] Hủy đơn → stock hoàn lại",
    "headers": TOKEN_HDR,
    "body": "",
    "pre": "Đơn hàng từ step 2, status=PENDING",
    "status": "200",
    "expected": "status=CANCELLED. SP id=9 stock hoàn về 55",
})

# ============================================================
# SAVE FILE
# ============================================================
out_path = r"e:\workspace_ptit\2-2025-2026\BTL-TMDT\be\context\testcases_laptopverse_be.xlsx"
wb.save(out_path)
print(f"✅ Excel saved: {out_path}")
print(f"   Sheets: {[s.title for s in wb.worksheets]}")
total = sum(ws.max_row - 1 for ws in wb.worksheets)
print(f"   Total rows (approx): {total}")
